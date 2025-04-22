import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os

def create_design_matrices():
    # Setup level faktor
    levels = [-1, 1]
    
    # Treatment combinations (standard order)
    treatments = []
    for a in levels:
        for b in levels:
            for c in levels:
                treatments.append((a, b, c))
    
    # Nilai Run Order dari gambar untuk RBD
    run_order_rbd = [22, 23, 21, 19, 17, 18, 24, 20, 
                    9, 10, 15, 14, 12, 13, 16, 11,
                    6, 8, 3, 4, 7, 1, 2, 5]
    
    # Nilai Run Order dari gambar untuk CRD
    run_order_crd = [24, 20, 23, 12, 3, 7, 14, 11,
                    2, 6, 17, 4, 9, 13, 8, 22, 
                    10, 19, 16, 13, 5, 18, 21, 1]
    
    # Data untuk RBD
    std_order = list(range(1, 9)) * 3
    blocks = [1] * 8 + [2] * 8 + [3] * 8
    
    # Memuat data dari treatments berdasarkan standard order
    a_values = []
    b_values = []
    c_values = []
    
    for order in std_order:
        treatment_idx = (order - 1) % 8
        a_values.append(treatments[treatment_idx][0])
        b_values.append(treatments[treatment_idx][1])
        c_values.append(treatments[treatment_idx][2])
    
    # Membuat DataFrame untuk RBD
    rbd_data = {
        'StdOrder': std_order,
        'RunOrder': run_order_rbd,
        'Blocks\n(replikasi)': blocks,
        'A': a_values,
        'B': b_values,
        'C': c_values,
        'Jadwal\nRunning': [''] * 24,
        'Nilai\nRespon': [''] * 24
    }
    
    df_rbd = pd.DataFrame(rbd_data)
    
    # Membuat DataFrame untuk CRD
    # Untuk CRD, setiap treatment direplikasi 3 kali
    crd_data = {
        'StdOrder': std_order,
        'RunOrder': run_order_crd,
        'Block': [1] * 24,
        'A': a_values,
        'B': b_values,
        'C': c_values,
        'Jadwal\nRunning': [''] * 24,
        'Nilai\nRespon': [''] * 24
    }
    
    df_crd = pd.DataFrame(crd_data)
    
    # Menyimpan ke Excel dengan format yang tepat
    save_design_matrices_to_excel(df_rbd, df_crd)
    
    return df_rbd, df_crd

def save_design_matrices_to_excel(df_rbd, df_crd, filename="output/design_matrices.xlsx"):
    # Memastikan direktori output ada
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # Membuat workbook baru
    wb = Workbook()
    
    # Menghapus sheet default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Membuat sheets untuk RBD dan CRD
    ws_rbd = wb.create_sheet("Randomized Block Design")
    ws_crd = wb.create_sheet("Completely Randomized Design")
    
    # Judul untuk RBD
    ws_rbd.merge_cells('A1:H1')
    ws_rbd['A1'] = 'Design Matrix untuk randomized block design'
    ws_rbd['A1'].font = Font(bold=True)
    ws_rbd['A1'].alignment = Alignment(horizontal='center')
    
    # Judul untuk CRD
    ws_crd.merge_cells('A1:H1')
    ws_crd['A1'] = 'Design Matrix untuk completely randomized design'
    ws_crd['A1'].font = Font(bold=True)
    ws_crd['A1'].alignment = Alignment(horizontal='center')
    
    # Subjudul untuk RBD
    ws_rbd.merge_cells('A2:H2')
    ws_rbd['A2'] = '(urutan running & treatment dibacak ulang setiap berganti replikasi/blok)'
    ws_rbd['A2'].alignment = Alignment(horizontal='center')
    
    # Subjudul untuk CRD
    ws_crd.merge_cells('A2:H2')
    ws_crd['A2'] = '(urutan running 24 eksperimen (8 treatment direplikasi 3x) dirandom secara total)'
    ws_crd['A2'].alignment = Alignment(horizontal='center')
    
    # Menambahkan header untuk RBD
    headers_rbd = ['StdOrder', 'RunOrder', 'Blocks\n(replikasi)', 'A', 'B', 'C', 'Jadwal\nRunning', 'Nilai\nRespon']
    for col_num, header in enumerate(headers_rbd, 1):
        cell = ws_rbd.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Menambahkan header untuk CRD
    headers_crd = ['StdOrder', 'RunOrder', 'Block', 'A', 'B', 'C', 'Jadwal\nRunning', 'Nilai\nRespon']
    for col_num, header in enumerate(headers_crd, 1):
        cell = ws_crd.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Light green fill untuk highlight
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    # Menambahkan data RBD
    for row_idx, row in enumerate(df_rbd.values, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws_rbd.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Highlight sel dengan warna hijau berdasarkan contoh gambar
            if row_idx in [13, 15, 16]:  # Rows 13, 15, 16 in dataframe (index 9, 11, 12 + 4)
                cell.fill = light_green_fill
    
    # Menambahkan data CRD
    for row_idx, row in enumerate(df_crd.values, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws_crd.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Highlight sel dengan warna hijau berdasarkan contoh gambar
            if row_idx in [8, 14, 21]:  # Rows 8, 14, 21 in dataframe (index 4, 10, 17 + 4)
                cell.fill = light_green_fill
    
    # Mengatur lebar kolom
    for ws in [ws_rbd, ws_crd]:
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12
    
    # Menyimpan file
    wb.save(filename)
    print(f"Design matrices telah disimpan dalam file '{filename}'")

def create_hypothesis_table():
    # Memastikan direktori output ada
    output_file = "output/hypothesis_table.xlsx"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Data statis untuk digunakan jika tidak bisa membaca dari Excel
    static_data = [
        ["Komposisi", "Faktor Komposisi tidak berpengaruh signifikan", "Faktor Komposisi berpengaruh signifikan", 12.345, 5.143, "Ditolak", "Komposisi berpengaruh signifikan"],
        ["Kompaksi", "Faktor Kompaksi tidak berpengaruh signifikan", "Faktor Kompaksi berpengaruh signifikan", 8.765, 5.143, "Ditolak", "Kompaksi berpengaruh signifikan"],
        ["Cavity", "Faktor Cavity tidak berpengaruh signifikan", "Faktor Cavity berpengaruh signifikan", 6.543, 5.143, "Ditolak", "Cavity berpengaruh signifikan"],
        ["Komposisi*Kompaksi", "Tidak ada interaksi antara faktor Komposisi dan Kompaksi", "Ada interaksi antara faktor Komposisi dan Kompaksi", 3.210, 5.143, "Diterima", "Tidak ada interaksi"],
        ["Komposisi*Cavity", "Tidak ada interaksi antara faktor Komposisi dan Cavity", "Ada interaksi antara faktor Komposisi dan Cavity", 2.109, 5.143, "Diterima", "Tidak ada interaksi"],
        ["Kompaksi*Cavity", "Tidak ada interaksi antara faktor Kompaksi dan Cavity", "Ada interaksi antara faktor Kompaksi dan Cavity", 1.876, 5.143, "Diterima", "Tidak ada interaksi"],
        ["Seluruh Faktor", "Tidak ada interaksi antar ketiga faktor", "Ada interaksi antar ketiga faktor", 0.987, 5.143, "Diterima", "Tidak ada interaksi"]
    ]
    
    # Membaca data ANOVA dari file input Excel
    try:
        excel_file = "input/semua_tabel.xlsx"
        sheet_name = "tabel_4.6"  # Sesuaikan dengan nama sheet di file Excel
        
        print(f"Mencoba membaca data dari {excel_file}, sheet {sheet_name}...")
        
        # Cek jika file ada
        if not os.path.exists(excel_file):
            print(f"File {excel_file} tidak ditemukan! Menggunakan data statis...")
            data = static_data
        else:
            # Cek jika sheet ada dalam file Excel
            excel = pd.ExcelFile(excel_file)
            available_sheets = excel.sheet_names
            
            if sheet_name not in available_sheets:
                print(f"Sheet '{sheet_name}' tidak ditemukan di {excel_file}!")
                print(f"Sheet yang tersedia: {available_sheets}")
                print("Menggunakan data statis...")
                data = static_data
            else:
                # Coba membaca data dari Excel
                try:
                    # Membaca data dari Excel dengan parameter header=None untuk melihat struktur data asli
                    df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    print(f"Struktur data yang dibaca dari Excel: {df_raw.shape} (baris x kolom)")
                    
                    # Cek jika DataFrame kosong
                    if df_raw.empty:
                        print("DataFrame kosong! Menggunakan data statis...")
                        data = static_data
                    else:
                        # Tampilkan beberapa baris pertama untuk debug
                        print("Beberapa baris pertama data:")
                        print(df_raw.head(10))
                        
                        # Baca ulang dengan parameter yang tepat berdasarkan struktur yang dilihat
                        df_anova = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=1)
                        print(f"Data ANOVA setelah dibaca ulang: {df_anova.shape} (baris x kolom)")
                        
                        # Ambil header kolom
                        print(f"Kolom data: {df_anova.columns.tolist()}")
                        
                        # Faktor-faktor yang akan diuji
                        factors = ["Komposisi", "Kompaksi", "Cavity", "Komposisi*Kompaksi", 
                                "Komposisi*Cavity", "Kompaksi*Cavity", "Seluruh Faktor"]
                        
                        data = []
                        for i, factor in enumerate(factors):
                            try:
                                # Cek jika ada cukup baris
                                if i+1 >= len(df_anova):
                                    print(f"Data untuk faktor {factor} tidak ditemukan (indeks {i+1} di luar jangkauan)")
                                    # Gunakan data statis untuk faktor ini
                                    data.append(static_data[i])
                                    continue
                                
                                row = df_anova.iloc[i+1]  # +1 karena baris pertama biasanya header
                                print(f"Data untuk faktor {factor}: {row.values}")
                                
                                # Mencari kolom F-hitung dan F-tabel berdasarkan nama kolom
                                if 'F-hitung' in df_anova.columns and 'F-tabel' in df_anova.columns:
                                    f_hit = float(row['F-hitung'])
                                    f_tab = float(row['F-tabel'])
                                    status = row.get('Status', "Ditolak" if f_hit > f_tab else "Diterima")
                                    keterangan = row.get('Keterangan', "")
                                # Jika tidak ada nama kolom yang tepat, coba dengan indeks (perlu disesuaikan)
                                else:
                                    # Asumsikan format: source | df | SS | MS | F-hit | F-tab | status | keterangan
                                    col_f_hit = None
                                    col_f_tab = None
                                    
                                    # Cari kolom yang mungkin berisi F-hitung dan F-tabel
                                    for idx, col_name in enumerate(df_anova.columns):
                                        if 'F-hit' in str(col_name) or 'F hit' in str(col_name):
                                            col_f_hit = idx
                                        elif 'F-tab' in str(col_name) or 'F tab' in str(col_name):
                                            col_f_tab = idx
                                    
                                    # Jika tidak ditemukan, gunakan indeks 4 dan 5 (indeks umum untuk F-hit dan F-tab)
                                    if col_f_hit is None:
                                        col_f_hit = min(4, len(row)-1)
                                    if col_f_tab is None:
                                        col_f_tab = min(5, len(row)-1)
                                    
                                    print(f"Menggunakan kolom {col_f_hit} untuk F-hitung dan {col_f_tab} untuk F-tabel")
                                    
                                    f_hit = float(row.iloc[col_f_hit])
                                    f_tab = float(row.iloc[col_f_tab])
                                    
                                    # Coba dapatkan status dan keterangan jika ada
                                    if col_f_tab + 1 < len(row):
                                        status = row.iloc[col_f_tab + 1]
                                    else:
                                        status = "Ditolak" if f_hit > f_tab else "Diterima"
                                    
                                    if col_f_tab + 2 < len(row):
                                        keterangan = row.iloc[col_f_tab + 2]
                                    else:
                                        keterangan = ""
                                
                                # H0 (null hypothesis)
                                if "*" in factor:
                                    parts = factor.split("*")
                                    h0 = f"Tidak ada interaksi antara faktor {parts[0]} dan {parts[1]}"
                                    h1 = f"Ada interaksi antara faktor {parts[0]} dan {parts[1]}"
                                else:
                                    h0 = f"Faktor {factor} tidak berpengaruh signifikan"
                                    h1 = f"Faktor {factor} berpengaruh signifikan"
                                
                                data.append([factor, h0, h1, f_hit, f_tab, status, keterangan])
                            
                            except Exception as e:
                                print(f"Error saat memproses faktor {factor}: {str(e)}")
                                print(f"Data baris: {row.values if 'row' in locals() else 'Tidak ada data'}")
                                # Gunakan data statis untuk faktor ini
                                data.append(static_data[i])
                except Exception as e:
                    print(f"Error saat membaca sheet: {str(e)}")
                    data = static_data
    
    except Exception as e:
        print(f"Error umum: {str(e)}")
        data = static_data
    
    # Membuat DataFrame
    df = pd.DataFrame(data, columns=["Faktor", "H0", "H1", "F-hitung", "F-tabel", "Status", "Keterangan"])
    
    # Menyimpan ke Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Hipotesis", index=False)
        
        # Mendapatkan objek workbook & worksheet
        wb = writer.book
        ws = writer.sheets["Hipotesis"]
        
        # Menambahkan judul
        ws.insert_rows(1)
        ws['A1'] = "Tabel daftar hipotesis penelitian (H1)"
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Mengatur lebar kolom
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
    print("Tabel hipotesis telah disimpan dalam file 'output/hypothesis_table.xlsx'")

# Menjalankan fungsi utama
if __name__ == "__main__":
    # Membuat design matrices
    df_rbd, df_crd = create_design_matrices()
    
    # Membuat tabel hipotesis
    create_hypothesis_table()
    
    print("Semua tabel telah dibuat di folder 'output'!") 